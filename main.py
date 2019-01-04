# -*- coding: utf-8 -*-
"""
本脚本实现从input文件夹的docx文件的相关数据
到output文件夹的excel表格

"""
import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')  # 编译环境utf8
from glob import glob
import re
import time
import string
try:
    from docx import Document
except ImportError:
    print(u'缺少模块python-docx，正在自动安装')
    import subprocess
    p = subprocess.Popen('pip install', shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    print(p.stdout.readlines())
    for line in p.stdout.readlines():
        print(line)
    retval = p.wait()
    from docx import Document
    # raise
try:
    import openpyxl
except ImportError:
    print(u'缺少模块openpyxl，正在自动安装')
    import subprocess
    p = subprocess.Popen('pip install openpyxl', shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    print(p.stdout.readlines())
    for line in p.stdout.readlines():
        print(line)
    retval = p.wait()
    import openpyxl
    # raise
##################################这是彩色打印
import ctypes
STD_INPUT_HANDLE = -10
STD_OUTPUT_HANDLE= -11
STD_ERROR_HANDLE = -12


FOREGROUND_BLACK = 0x0
FOREGROUND_BLUE = 0x01 # text color contains blue.
FOREGROUND_GREEN= 0x02 # text color contains green.
FOREGROUND_RED = 0x04 # text color contains red.
FOREGROUND_INTENSITY = 0x08 # text color is intensified.

BACKGROUND_BLUE = 0x10 # background color contains blue.
BACKGROUND_GREEN= 0x20 # background color contains green.
BACKGROUND_RED = 0x40 # background color contains red.
BACKGROUND_INTENSITY = 0x80 # background color is intensified.
#上面这一大段都是在设置前景色和背景色，其实可以用数字直接设置，我的代码直接用数字设置颜色


class Color:
    std_out_handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)

    def set_cmd_color(self, color, handle=std_out_handle):
        bool = ctypes.windll.kernel32.SetConsoleTextAttribute(handle, color)
        return bool

    def reset_color(self):
        self.set_cmd_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE | FOREGROUND_INTENSITY)
        #初始化颜色为黑色背景，纯白色字，CMD默认是灰色字体的

    def print_red_text(self, print_text):
        self.set_cmd_color(4 | 8)
        print(print_text)
        self.reset_color()
        #红色字体

    def print_green_text(self, print_text):
        self.set_cmd_color(FOREGROUND_GREEN | FOREGROUND_INTENSITY)
        # c = raw_input(print_text.encode('gbk'))
        # c = raw_input(print_text)
        print(print_text)
        self.reset_color()
        # return c

    def print_yellow_text(self, print_text):
        self.set_cmd_color(6 | 8)
        print(print_text)
        self.reset_color()
        #黄色字体

    def print_blue_text(self, print_text):
        self.set_cmd_color(1 | 10)
        print(print_text)
        self.reset_color()
        #蓝色字体


clr = Color()
clr.set_cmd_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE | FOREGROUND_INTENSITY)
# clr.print_red_text('red')
# clr.print_green_text("green")
# clr.print_blue_text('blue')
# clr.print_yellow_text('yellow')
##########################################


PROJECT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))


re_pats = {
    'date': u'\d+年\d+月\d+日',
    'stage': u'【(\d+)】',
    'money': u'人民币(.*?)元',
}


def get_filename_by_path(path, forbid_word=''):
    searched_filenames = glob(path)
    return [i for i in searched_filenames if forbid_word not in i ]

def parse_text_by_repat(text, re_pat):
    find = re.findall(re_pat, text)
    return (find[0].strip() if find else None)


def read_docx(filename):
    results = []
    document = Document(filename)
    print(u'定位并解析数据')
    fillcontent = False
    content = ""
    
    for para_num, para in enumerate(document.paragraphs):
        try:
            debug = True
            if para.style.name.startswith("Heading"):
                result = {
                    'FileName': None,
                    'Title': None,
                    'ReqId': None,
                    'Content': None,
                }
                #print('line %d' % para_num)
                style = para.style
                format = style.paragraph_format
                if debug:
                    print(" text:%s" % para.text)
                    print(" style:%s" % style.name)
                
                while style.next_paragraph_style != style:
                    style = style.next_paragraph_style
                    if debug:
                        print(" style:%s" % style.name)

                for run_num, run in enumerate(para.runs):
                    style = para.style
                    format = style.paragraph_format
                    if debug:
                        print("  text:%s" % run.text)
                        print("   style:%s" % run.style.name)
                        while style.next_paragraph_style != style:
                            style = style.next_paragraph_style
                            if debug:
                                print("   style:%s" % style.name)
                        print("   strike:%s" % run.font.strike)

                    if run.font.strike != True and len(run.text) > 1:
                        #if run.style.name.startswith("Default Paragraph Font"):
                        #    if result["Title"] == None:
                        if result["Title"] == None:
                            result["Title"] = run.text
                        #else:
                        #    if result["ReqId"] == None:
                        elif result["ReqId"] == None:
                            result["ReqId"] = run.text

                if result["Title"] != None and  result["ReqId"] != None:
                    fillcontent = True
                    print("Title:%s" % result["Title"])
                    print("ReqId:%s" % result["ReqId"])  
                    print(" ")
                    result["FileName"] = os.path.basename(filename)
                    if len(results) > 1: # skip the first one
                        results[-1]["Content"] = content # save the last one content
                    content = "" # empty the content
                    results.append(result)

            else:
                if len(para.text) > 1:
                    content = content + para.text + "\n"
                #    #print(para.text)
                #    print(para.style.name)
        except UnicodeEncodeError:
            print(" UnicodeEncodeError")
     
    #save the last para
    if len(results) > 1: # skip the first one
        results[-1]["Content"] = content # save the last one content
    content = "" # empty the content
        
    print("count %d" % len(results))
    return results

def write_excel(excel_name, result_dicts):
    from openpyxl.workbook import Workbook
    
    from openpyxl.styles import Alignment
    alignment = Alignment(
        wrap_text = True, # 自动换行
    )

    #ExcelWriter,里面封装好了对Excel的写操作
    from openpyxl.writer.excel import ExcelWriter

    #get_column_letter函数将数字转换为相应的字母，如1-->A,2-->B
    from openpyxl.utils  import get_column_letter

    from openpyxl.reader.excel import load_workbook

    if os.path.isfile(excel_name):
        # #读取excel2007文件
        wb = load_workbook(excel_name)
    else:
        #新建一个workbook
        wb = Workbook()

    #设置文件输出路径与名称
    dest_filename = excel_name

    
    # # 获取第一个sheet

    ws = wb.get_active_sheet()
    if ws != None:
        wb.remove_sheet(ws)
       
    ws = wb.create_sheet('Sheet1')


    #第一个sheet是ws
    # ws = wb.worksheets[0]

    # #设置ws的名称
    # ws.title = "sheet1"

    line = 1
    print(u'定位写入坐标')
    while ws.cell(line,1).value:
        # print(ws.cell("A%s" % line).value)
        line += 1
    print(u'从第%s行开始写入' % line)


    #Title
    ws.cell(line,1).value=u'FileName'
    ws.cell(line,2).value=u'Title'
    ws.cell(line,3).value=u'ReqId'
    ws.cell(line,4).value=u'Content'
    ws.cell(line,5).value=u'Owner'
    ws.column_dimensions['D'].width = 50.0
    line += 1
    
    for i, result in enumerate(result_dicts):
        print(u'正在写入第%s条数据到excel' % (i+1))
        print(u'正在写入 %s' % result['FileName'])
        print(u'正在写入 %s' % result['Title'])
        print(u'正在写入 %s' % result['ReqId'])
        ws.cell(line,1).value=result['FileName']
        ws.cell(line,2).value=result['Title']
        ws.cell(line,3).value=result['ReqId']
        ws.cell(line,4).value=result['Content']
        ws.cell(line,4).alignment=alignment
        line += 1

    #最后保存文件
    wb.save(filename=dest_filename)
    
def main():
    print(u'开始执行')
    print(u'从input文件夹查找docx文件')
    filenames = get_filename_by_path('input/*.docx', '~$')
    result_dicts = []
    for filename in filenames:
        print(u'读取文件：')
        clr.print_blue_text(os.path.basename(filename))
        
        results = read_docx(filename)
        #print(len(results))
        #add for each one
        for result in results:
            result_dicts.append(result)
    save_filename = 'output/output.xlsx'
    # save_filename = 'output/output%s.xlsx' % int(time.time())
    write_excel(save_filename, result_dicts)
    print(u'执行完毕，文件保存至')
    clr.print_blue_text(save_filename)
    # print(save_filename)
    print(u'敲击回车结束运行')
    raw_input()

if __name__ == '__main__':
    main()
